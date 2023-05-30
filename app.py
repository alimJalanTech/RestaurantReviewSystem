import streamlit as st
import pandas as pd
import plotly.express as px
import openpyxl as op
import webbrowser
from pathlib import Path
import pickle
import streamlit_authenticator as stauth
from collections import defaultdict
from PIL import Image

# Set page configuration
st.set_page_config(page_title="Customer Reviews")

names = ["Alim", "Sahil", "Devashish"]
usernames = ["ma", "msk", "dk"]
password = ["AM", "KSM", "KD"]

# load hashed passwords
file_path = Path(__file__).parent / "hashed_pw.pkl"
with file_path.open("rb") as file:
    hashed_passwords = pickle.load(file)

credentials = {"usernames": {}}

for un, name, pw in zip(usernames, names, hashed_passwords):
    user_dict = {"name": name, "password": pw}
    credentials["usernames"].update({un: user_dict})

authenticator = stauth.Authenticate(
    credentials, "Customer Reviews", "abcdef", 30)
name, authentication, username = authenticator.login("Login", "main")


if authentication == False:
    st.error("Username/password is incorrect")

if authentication == None:
    st.warning("Please enter your username and password")

if authentication:
    with open("style.css") as f:
        st.markdown("<style>{}</style>".format(f.read()),
                    unsafe_allow_html=True)

    # Wrap all content in a container div
    st.markdown(
        """
        <div style='display: flex; flex-direction: column; align-items: center;'>
        """, unsafe_allow_html=True)

    st.header("Analyzed Customer Reviews")
    head_image = Image.open('headimage.jpg')
    # # Reduce image size
    # desired_size = (800, 600)  # Specify your desired size
    # resized_image = head_image.resize(desired_size)
    st.image(head_image, caption='Our Restaurant', use_column_width=True)
    


    excel_file = 'Reviews_Ans.xlsx'
    excel_file2 = 'opinion.xlsx'
    sheet = 'Sheet1'


    # Takes the following total columns and presents them
    st.header("Customer Information")
    df = pd.read_excel(excel_file,
                       sheet_name=sheet,
                       usecols='A:D',
                       header=0,)
    authenticator.logout("Logout", "sidebar")
    st.sidebar.title(f"Welcome {name}")
    
    dishbook_op = op.load_workbook('Reviews_Ans.xlsx')
    dish_op = dishbook_op.active
    
    #New Code for Dishes
    foods = defaultdict(int)
    i = 1
    while True:
        good = dish_op.cell(row=i, column=4).value
        if good == "NULL" or good == "FoodItem":
            i += 1 
            continue
        if good == None:
            break
        foods[good] += 1
        i += 1
    dishbook_op.save('Reviews_Ans.xlsx')

    fd = pd.DataFrame.from_dict(foods, orient='index', columns=['Customers'])
    fd.reset_index(inplace=True)
    fd.rename(columns={'index': 'Dish'}, inplace=True)

    # Create the bar graph using Plotly Express
    gih = px.bar(fd, x='Dish', y='Customers', title='Number of Customers per Dish')

    
    # Take the opinion excel file and generates a pie chart
    opinion = pd.read_excel(excel_file2,
                            sheet_name=sheet,
                            usecols='A:C',
                            header=0,)
    workbook_op = op.load_workbook('opinion.xlsx')
    sheet_op = workbook_op.active
    good = sheet_op.cell(row=2, column=1).value
    bad = sheet_op.cell(row=2, column=2).value

    # Create pie
    pie = px.pie(title="Reviews", values=[good, bad], names=["Good", "Bad"])

    workbook_op.save('opinion.xlsx')

    link = "https://t.me/myRestaurant123_bot"
    button_label = "Go to Telegram"

    if st.button(button_label):
        webbrowser.open_new_tab(link)

    st.image("telegram.png", "Send a message to users", 100)

    st.dataframe(df)
    st.dataframe(opinion)

    st.plotly_chart(pie)

    st.subheader("Customers who bought what food")
    st.write(foods)
    st.plotly_chart(gih)


    # Visualize the histogram

    # Close the container div
    st.markdown("</div>", unsafe_allow_html=True)
